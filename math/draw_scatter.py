#!/usr/bin/env python3
# -*- coding: UTF-8 -*-

'''

@File   : draw_scatter.py   
@Author : alexander.here@gmail.com   
@Date   : 2020-07-07 15:29 CST(+0800)   
@Brief  :  

'''

import os, sys
import csv
import numpy as np
import warnings
import matplotlib.pyplot as plt
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    DND_ENABLED = True
except:
    DND_ENABLED = False
import tkinter as tk
import tkinter.scrolledtext as tkst
import tkinter.messagebox as tkmb
from tkinter import ttk
from tkinter import filedialog as tkfd
try:
    import openpyxl
    openpyxl_loaded = True
except:
    openpyxl_loaded = False
try:
    import xlrd
    xlrd_loaded = True
except:
    xlrd_loaded = False

# Defines:

DIR_VERTICAL = 0
DIR_HORIZONTAL = 1

TL_NONE = 'NONE'
TL_AVG = 'Moving Average'
TL_POLY = 'Polynomial'
TL_VERTICAL = 'Vertical'
TL_HORIZONTAL = 'Horizontal'

COLORS = [
    '#069af3', # azure
    '#ef4026', # tomato
    '#029386', # teal
    '#fac205', # goldenrod
    '#580f41', # plum
    '#bbf90f', # yellowgreen
    '#00ffff', # cyan
    '#ff81c0', # pink
    '#380282', # indigo
    '#a9561e', # sienna
]

MARKERS = [ '^', 'x', 'd', '+', 'o', 'v', 's', '<', '1', '*', '>',]

default_file = None
try:
    if os.path.isfile( sys.argv[ 1]):
        default_file = sys.argv[ 1]
except:
    pass

# Basic Functions:

def choose_marker( idx):
    return MARKERS[ idx % len( MARKERS)]

def choose_color( idx):
    if idx < len( COLORS):
        return COLORS[ idx]
    hex = [ '0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'a', 'b', 'c', 'd', 'e', 'f']
    color = '#'
    color += hex[ ( 2 + idx * 13) % 16]
    color += hex[ ( 3 + idx * 1) % 16]
    color += hex[ ( 2 + idx * 7) % 16]
    color += hex[ ( 4 + idx * 3) % 16]
    color += hex[ ( 2 + idx * 3) % 16]
    color += hex[ ( 5 + idx * 5) % 16]
    return color

sheet_cache = dict()

def load_sheet( path, sheet):
    if len( path) <= 0:
        raise Exception( "Missing Data file!")
    if not os.path.isfile( path):
        raise Exception( "File '" + path + "' NOT exists!")
    if path.lower().endswith( '.csv') or path.lower().endswith( '.txt'):
        return load_csv_sheet( path)
    elif path.lower().endswith( '.xls'):
        return load_old_excel_sheet( path, sheet)
    elif path.lower()[-5:] in { '.xlsx', '.xlsm', '.xltx', 'xltm'}:
        return load_excel_sheet( path, sheet)
    else:
        raise Exception( "Do NOT support type of file '" + path + "'!")

def load_csv_sheet( path):
    abs_path = os.path.abspath( path)
    if abs_path in sheet_cache:
        return sheet_cache[ abs_path]
    try:
        with open( path, newline = '') as csv_file:
            dialect = csv.Sniffer().sniff( csv_file.read())
            csv_file.seek( 0)
            reader = csv.reader( csv_file, dialect)
            sheet = []
            for row in reader:
                sheet.append( list( row))
            sheet_cache[ abs_path] = sheet
            return sheet
    except Exception as e:
        raise Exception( "Failed parsing CSV file: " + str( e))

def load_old_excel_sheet( path, sheet_name):
    if not xlrd_loaded:
        raise Exception( "need module 'xlrd' to load data from file '" + path + "'!")
    sheet_key = ( os.path.abspath( path), sheet_name)
    if sheet_key in sheet_cache:
        return sheet_cache[ sheet_key]
    try:
        wb = xlrd.open_workbook( path)
        if not sheet_name or len( sheet_name) <= 0:
            sheet = wb.sheet_by_index( 0)
        elif sheet_name in wb.sheet_names():
            sheet = wb.sheet_by_name( sheet_name)
        else:
            try:
                sheet = wb.sheet_by_index( int( sheet_name))
            except:
                raise Exception( "invalid sheet name '" + sheet_name + "'")
        result = []
        for row_id in range( sheet.nrows):
            result.append( [ str( value) for value in sheet.row_values( row_id)])
        sheet_cache[ sheet_key] = result
        return result
    except Exception as e:
        raise Exception( "Failed loading Excel file: " + str( e))

def load_excel_sheet( path, sheet_name):
    if not openpyxl_loaded:
        raise Exception( "need module 'openpyxl' to load data from file '" + path + "'!")
    sheet_key = ( os.path.abspath( path), sheet_name)
    if sheet_key in sheet_cache:
        return sheet_cache[ sheet_key]
    try:
        wb = openpyxl.load_workbook( path, read_only = True, data_only = True)
        if not sheet_name or len( sheet_name) <= 0:
            sheet = wb[ wb.sheetnames[ 0]]
        elif sheet_name in wb.sheetnames:
            sheet = wb[ sheet_name]
        else:
            try:
                sheet = wb[ wb.sheetnames[ int( sheet_name)]]
            except:
                raise Exception( "invalid sheet name '" + sheet_name + "'")
        result = []
        for row in sheet.rows:
            row_data = [ cell.value for cell in row]
            for col in range( len( row_data)):
                if row_data[ col] is None:
                    row_data[ col] = ''
                elif type( row_data[ col]) is str:
                    pass
                else:
                    row_data[ col] = str( row_data[ col])
            result.append( row_data)
        # remove extra rows:
        while len( result) > 0:
            all_none = True
            for v in result[ -1]:
                if v != '':
                    all_none = False
                    break
            if all_none:
                del result[ -1]
            else:
                break
        sheet_cache[ sheet_key] = result
        return result
    except Exception as e:
        raise Exception( "Failed loading Excel file: " + str( e))

def parse_sheet_index( s):
    s = s.strip()
    if s == '':
        return -1
    try:
        return int( s) - 1
    except:
        pass
    result = 0
    for ch in s.upper():
        v = ord( ch) - ord( 'A') + 1
        if v < 1 or v > 26:
            raise Exception( "Not valid index '" + s + "'!")
        result = result * 26 + v
    return result - 1

def parse_number( s):
    s = s.strip()
    try:
        return int( s)
    except:
        pass
    try:
        return float( s)
    except:
        raise Exception( "'" + s + "' is NOT a number!")

def parse_filter_line( line, sheet, direction, idx):
    result = ''
    while True:
        start = line.find( '${')
        if start < 0:
            result += line
            return result
        length = line[ start+2:].find( '}')
        if length < 0:
            raise Exception( "brace('{}') NOT match!")
        parsed_idx = parse_sheet_index( line[ start+2:start+2+length])
        if direction == DIR_HORIZONTAL:
            row, col = parsed_idx, idx
        elif direction == DIR_VERTICAL:            
            row, col = idx, parsed_idx
        else:
            raise Exception( "Internal Exception 78D2!")
        if row < len( sheet) and col < len( sheet[ row]):
            value = "'" + sheet[ row][ col] + "'"
        else:
            value = "''"
        result += line[ :start]
        result += value
        line = line[ start+3+length:]

def check_filters( filters, sheet, direction, idx):
    if filters is None or len( filters) <= 0:
        return True
    try:
        exp = parse_filter_line( ' '.join( filters.splitlines()), sheet, direction, idx).strip()
        if len( exp) > 0:
            value = eval( exp)
            if not value or value == '':
                return False
    except Exception as e:
        raise Exception( "Invalid filter: " + str( e))
    return True

def generate_trend( x, y, t, n):
    if t == TL_HORIZONTAL:
        try:
            n = float( n)
        except:
            raise Exception( 'Invalid trendline param: N = ' + str( n) + ' (number expected)')
        yy = [ n, n]
        xx = [ min( x), max( x)]
        return xx, yy
    elif t == TL_VERTICAL:
        try:
            n = float( n)
        except:
            raise Exception( 'Invalid trendline param: N = ' + str( n) + ' (number expected)')
        xx = [ n, n]
        yy = [ min( y), max( y)]
        return xx, yy
    elif t == TL_AVG:
        try:
            n = int( n)
            if n < 2:
                raise Exception()
        except:
            raise Exception( 'Invalid trendline param: N = ' + str( n) + ' (N>1 expected)')
        count = len( x)
        data = sorted( zip( x, y), key = lambda d:d[0])
        x = [ d[ 0] for d in data]
        y = [ d[ 1] for d in data]
        xx = []
        yy = []
        last_start = -n
        for i in range( 1 - n, count):
            start = i
            while start > 0 and x[ start] == x[ start-1]:
                start -= 1
            if start == last_start:
                continue
            end = min( start + n - 1, count - 1)
            while end < count - 1 and x[ end] == x[ end+1]:
                end += 1
            start = max( 0, start)
            xx.append( ( x[ start] + x[ end]) / 2.0)
            yy.append( sum( y[ start:end+1]) / float( end + 1 - start))
        return xx, yy
    elif t == TL_POLY:
        try:
            n = int( n)
            if n < 0 or n > 5:
                raise Exception()
        except:
            raise Exception( 'Invalid trendline param: N = ' + str( n) + ' (0<N<6 expected)')
        poly = np.polyfit( x, y, n)
        f = np.poly1d( poly)
        i = min( x)
        end = max( x)
        step = ( end - i) / 20.0
        xx = []
        while i <= end:
            xx.append( i)
            i += step
        return xx, f( xx)
    return None, None

# Init:

if DND_ENABLED:
    window_main = TkinterDnD.Tk()
else:
    window_main = tk.Tk()
window_main.title( 'Scatter Config')

var_global_title = tk.StringVar()
var_global_tltype = tk.StringVar()
var_global_tln = tk.StringVar()
var_dir = tk.IntVar()
var_dir.set( 0)
var_name = tk.StringVar()
var_file = tk.StringVar()
var_x = tk.StringVar()
var_y = tk.StringVar()
var_from = tk.StringVar()
var_to = tk.StringVar()
var_sheet = tk.StringVar()
var_tltype = tk.StringVar()
var_tln = tk.StringVar()
var_hide = tk.IntVar()

plt.rcParams[ 'axes.unicode_minus'] = False
plt.rcParams[ 'font.sans-serif'] = [ 'SimHei', 'Heiti TC', 'Adobe Heiti Std', 'Adobe Fan Heiti Std']
warnings.simplefilter( 'ignore', np.RankWarning)

# Core:

class Serial:
    def __init__( self, name = None):
        self.name = name
        self.file = ''
        self.dir = DIR_VERTICAL
        self.from_ = ''
        self.to = ''
        self.x = ''
        self.y = ''
        self.sheet = ''
        self.trend = TL_NONE
        self.n = 0
        self.hide = 0
        self.filters = ''
    def set( self, other):
        self.name       = other.name
        self.file       = other.file
        self.dir        = other.dir
        self.from_      = other.from_
        self.to         = other.to
        self.x          = other.x
        self.y          = other.y
        self.sheet      = other.sheet
        self.trend      = other.trend
        self.n          = other.n
        self.hide       = other.hide
        self.filters    = other.filters
    def __eq__( self, other):
        return isinstance( other, Serial) \
            and self.name       == other.name \
            and self.file       == other.file \
            and self.dir        == other.dir \
            and self.from_      == other.from_ \
            and self.to         == other.to \
            and self.x          == other.x \
            and self.y          == other.y \
            and self.sheet      == other.sheet \
            and self.trend      == other.trend \
            and self.n          == other.n \
            and self.hide       == other.hide \
            and self.filters    == other.filters \
    
class SerialManager:
    serial_list = list()
    loaded_pos = -1
    new_group_id = 1
    changed = True
    def add( name = None):
        if name is None:
            name = "Group " + str( SerialManager.new_group_id)
        SerialManager.new_group_id += 1
        SerialManager.changed = True
        pos = len( SerialManager.serial_list)
        SerialManager.serial_list.append( Serial( name))
        if len( SerialManager.serial_list) > 1:
            SerialManager.serial_list[ -1].file = SerialManager.serial_list[ -2].file
            SerialManager.serial_list[ -1].from_ = SerialManager.serial_list[ -2].from_
            SerialManager.serial_list[ -1].to = SerialManager.serial_list[ -2].to
            SerialManager.serial_list[ -1].dir = SerialManager.serial_list[ -2].dir
            SerialManager.serial_list[ -1].x = SerialManager.serial_list[ -2].x
            SerialManager.serial_list[ -1].y = SerialManager.serial_list[ -2].y
            SerialManager.serial_list[ -1].sheet = SerialManager.serial_list[ -2].sheet
            SerialManager.serial_list[ -1].trend = SerialManager.serial_list[ -2].trend
            SerialManager.serial_list[ -1].n = SerialManager.serial_list[ -2].n
            SerialManager.serial_list[ -1].hide = SerialManager.serial_list[ -2].hide
            SerialManager.serial_list[ -1].filters = SerialManager.serial_list[ -2].filters
        if default_file and ( not SerialManager.serial_list[ -1].file or len( SerialManager.serial_list[ -1].file) <= 0):
            SerialManager.serial_list[ -1].file = default_file
        return pos
    def remove( idx):
        if idx >= 0 and idx < len( SerialManager.serial_list):
            SerialManager.changed = True
            del SerialManager.serial_list[ idx]
            if idx == SerialManager.loaded_pos:
                SerialManager.loaded_pos = -1
    def name( idx = None):
        if idx is None:
            idx = SerialManager.loaded_pos
        if idx >= 0 and idx < len( SerialManager.serial_list):
            return SerialManager.serial_list[ idx].name
        return ''
    def display_name( idx = None):
        if idx is None:
            idx = SerialManager.loaded_pos
        if idx >= 0 and idx < len( SerialManager.serial_list):
            if SerialManager.serial_list[ idx].hide > 0:
                return '# ' + SerialManager.serial_list[ idx].name
            else:
                return SerialManager.serial_list[ idx].name
        return ''
    def get( idx = None):
        if idx is None:
            idx = SerialManager.loaded_pos
        if idx >= 0 and idx < len( SerialManager.serial_list):
            return SerialManager.serial_list[ idx]
        return None
    def load( idx):
        ret = SerialManager.get( idx)
        if ret is not None:
            SerialManager.loaded_pos = idx
        return ret
    def save( data, idx = None):
        if idx is None:
            idx = SerialManager.loaded_pos
        if idx >= 0 and idx < len( SerialManager.serial_list):
            if SerialManager.serial_list[ idx] != data:
                SerialManager.changed = True
                SerialManager.serial_list[ idx].set( data)
    def count():
        return len( SerialManager.serial_list)
    def id_range():
        return range( len( SerialManager.serial_list))
    def title():
        if len( SerialManager.serial_list) > 0:
            ret = None
            for s in SerialManager.serial_list:
                name = os.path.basename( s.file)
                if ret is None:
                    ret = name
                elif ret != name:
                    return 'Scatter'
            return ret
        return ''
    def draw():
        if len( SerialManager.serial_list) <= 0 or not SerialManager.changed:
            return
        try :
            plt.cla()
            all_x = []
            all_y = []
            for i in SerialManager.id_range():
                name = SerialManager.serial_list[ i].name
                # basic info:
                x_idx = parse_sheet_index( SerialManager.serial_list[ i].x)
                if x_idx < 0:
                    raise Exception( 'Missing index of X values!')
                y_idx = parse_sheet_index( SerialManager.serial_list[ i].y)
                if x_idx < 0:
                    raise Exception( 'Missing index of Y values!')
                from_idx = parse_sheet_index( SerialManager.serial_list[ i].from_)
                to_idx = parse_sheet_index( SerialManager.serial_list[ i].to)
                sheet = load_sheet( SerialManager.serial_list[ i].file, SerialManager.serial_list[ i].sheet)
                row_count = len( sheet)
                col_count = 0
                for row in sheet:
                    l = len( row)
                    if l > col_count:
                        col_count = l
                color = choose_color( i)
                marker = choose_marker( i)
                # gather data:
                x = []
                y = []
                if SerialManager.serial_list[ i].dir == DIR_VERTICAL:
                    if from_idx < 0:
                        range_start = 0
                    else:
                        range_start = from_idx
                    if to_idx < 0:
                        range_end = row_count
                    else:
                        range_end = to_idx + 1
                    if range_start > range_end:
                        range_start, range_end = range_end - 1, range_start + 1
                    for row in range( range_start, range_end):
                        if row >= row_count or x_idx >= len( sheet[ row]) or y_idx >= len( sheet[ row]) \
                                or sheet[ row][ x_idx].strip() == '' or sheet[ row][ y_idx].strip() == '':
                            continue
                        if not check_filters( SerialManager.serial_list[ i].filters, sheet, DIR_VERTICAL, row):
                            continue
                        x.append( parse_number( sheet[ row][ x_idx]))
                        y.append( parse_number( sheet[ row][ y_idx]))
                elif SerialManager.serial_list[ i].dir == DIR_HORIZONTAL:
                    if from_idx < 0:
                        range_start = 0
                    else:
                        range_start = from_idx
                    if to_idx < 0:
                        range_end = col_count
                    else:
                        range_end = to_idx + 1
                    if range_start > range_end:
                        range_start, range_end = range_end - 1, range_start + 1
                    for col in range( range_start, range_end):
                        if x_idx >= row_count or y_idx >= row_count or col >= len( sheet[ x_idx]) or col >= len( sheet[ y_idx]) \
                                or sheet[ x_idx][ col].strip() == '' or sheet[ y_idx][ col].strip() == '':
                            continue
                        if not check_filters( SerialManager.serial_list[ i].filters, sheet, DIR_HORIZONTAL, col):
                            continue
                        x.append( parse_number( sheet[ x_idx][ col]))
                        y.append( parse_number( sheet[ y_idx][ col]))
                else:
                    raise Exception( "Internal Exception F676")
                if len( x) <= 0 or len( y) <= 0:
                    continue
                if var_global_tltype.get() != TL_NONE:
                    all_x += x
                    all_y += y
                if SerialManager.serial_list[ i].hide:
                    continue
                # draw scatter:
                plt.scatter( x, y, c = color, marker = marker, label = name)
                # draw trend line:
                trend_x, trend_y = generate_trend( x, y, SerialManager.serial_list[ i].trend, SerialManager.serial_list[ i].n)
                if trend_x is not None and trend_y is not None:
                    plt.plot( trend_x, trend_y, color = color + '48')
            # global trend:
            trend_x, trend_y = generate_trend( all_x, all_y, var_global_tltype.get(), var_global_tln.get())
            if trend_x is not None and trend_y is not None:
                plt.plot( trend_x, trend_y, color = '#00000010')
        except Exception as e:
            raise Exception( "Failed drawing '" + name + "': " + str( e))
        finally:
            SerialManager.changed = False
        sheet_cache.clear()
        if len( var_global_title.get().strip()) > 0:
            plt.title( var_global_title.get().strip())
        else:
            plt.title( SerialManager.title())
        plt.grid()
        plt.legend()
        plt.draw()
        plt.pause(0.01)



# Callbacks:

def print_event( evt = None):
    print( str( evt))

def do_nothing( evt = None):
    pass

def do_exit( evt = None):
    exit( 0)

def refresh_serial_list():
    global listbox_serial
    listbox_serial.delete( 0, tk.END)
    for i in SerialManager.id_range():
        listbox_serial.insert( tk.END, SerialManager.display_name( i))

def highlight_serial( pos):
    global listbox_serial
    if pos < SerialManager.count():
        listbox_serial.see( pos)
        listbox_serial.activate( pos)
        listbox_serial.select_set( pos)
        serial_2_ui( SerialManager.load( pos))

def serial_2_ui( serial = None):
    if serial is None:
        var_name.set( '')
        var_file.set( '')
        var_dir.set( 0)
        var_from.set( '')
        var_to.set( '')
        var_x.set( '')
        var_y.set( '')
        var_sheet.set( '')
        var_tltype.set( '')
        var_tln.set( '')
        text_filter.delete( '1.0', tk.END)
    else:
        var_name.set( serial.name)
        var_file.set( serial.file)
        var_dir.set( serial.dir)
        var_from.set( serial.from_)
        var_to.set( serial.to)
        var_x.set( serial.x)
        var_y.set( serial.y)
        var_sheet.set( serial.sheet)
        var_tltype.set( serial.trend)
        var_tln.set( serial.n)
        var_hide.set( serial.hide)
        text_filter.delete( '1.0', tk.END)
        text_filter.insert( tk.END, serial.filters)

def ui_2_serial():
    serial = Serial( var_name.get())
    serial.file = var_file.get()
    serial.dir = var_dir.get()
    serial.from_ = var_from.get()
    serial.to = var_to.get()
    serial.x = var_x.get()
    serial.y = var_y.get()
    serial.sheet = var_sheet.get()
    serial.trend = var_tltype.get()
    serial.n = var_tln.get()
    serial.hide = var_hide.get()
    serial.filters = text_filter.get( '1.0', tk.END)[:-1]
    return serial

def save_current_serial():
    SerialManager.save( ui_2_serial())
    pos = SerialManager.loaded_pos
    listbox_serial.delete( pos)
    listbox_serial.insert( pos, SerialManager.display_name())
    try:
        SerialManager.draw()
    except Exception as e:
        tkmb.showerror( title = 'ERROR', message = str( e))

def click_add():
    save_current_serial()
    pos = SerialManager.add()
    refresh_serial_list()
    highlight_serial( pos)

def click_del():
    global listbox_serial
    sels = listbox_serial.curselection()
    if len( sels) > 0:
        for sel in sels:
            SerialManager.remove( sel)
        refresh_serial_list()
        pos = min( sels) - 1
        highlight_serial( pos)
    try:
        SerialManager.draw()
    except Exception as e:
        tkmb.showerror( title = 'ERROR', message = str( e))


def click_serial( evt = None):
    sels = listbox_serial.curselection()
    if len( sels) == 1 and sels[0] != SerialManager.loaded_pos:
        save_current_serial()
        serial_2_ui( SerialManager.load( sels[ 0]))

def double_click_serial( evt = None):
    sels = listbox_serial.curselection()
    if len( sels) == 1:
        var_hide.set( not var_hide.get())
        click_update()

def click_update( evt = None):
    SerialManager.changed = True
    save_current_serial()

def click_browse( evt = None):
    types = [ 'csv']
    if xlrd_loaded:
        types += [ '.xls']
    if openpyxl_loaded:
        types += [ '.xlsx', '.xlsm', '.xltx', '.xltm']
    path = tkfd.askopenfilename( title = 'Select data source', filetypes = [('Spreadsheets', ' '.join( types)), ('ALL', '*.*')])
    if len( path) > 0 and os.path.isfile( path):
        if SerialManager.loaded_pos < 0:
            global default_file
            default_file = path
            click_add()
        else:
            var_file.set( path)

def on_dnd_enter( evt):
    evt.widget.focus_force()
    return evt.action

def on_dnd_drop( evt):
    if evt.data:
        path = window_main.tk.splitlist( evt.data)[ 0]
        if os.path.isfile( path):
            global default_file
            default_file = path
            click_add()

# Widgets & Layout:

frame_global = tk.Frame( window_main)
frame_global.pack( side = tk.TOP, fill = tk.X, padx = 8, pady = 4)

tk.Label( frame_global, anchor = tk.W, text = "TITLE").pack( side = tk.LEFT, padx = 4)

entry_global_title = tk.Entry( frame_global, textvariable = var_global_title)
entry_global_title.pack( side = tk.LEFT, fill = tk.X, expand = True)

tk.Label( frame_global, anchor = tk.W, text = "Total Trend:").pack( side = tk.LEFT, padx = 4)
combobox_global_tltype = ttk.Combobox( frame_global, textvariable = var_global_tltype, state = 'readonly', values = ( TL_NONE, TL_POLY, TL_AVG, TL_VERTICAL, TL_HORIZONTAL))
combobox_global_tltype.current( 0)
combobox_global_tltype.pack( side = tk.LEFT)

tk.Label( frame_global, anchor = tk.W, text = "N:").pack( side = tk.LEFT)
entry_global_n = tk.Entry( frame_global, textvariable = var_global_tln, width = 4)
entry_global_n.pack( side = tk.LEFT)

frame_serial = tk.Frame( window_main)
frame_serial.pack( fill = tk.BOTH, expand = True, padx = 4, pady = 4)

listbox_serial = tk.Listbox( frame_serial, borderwidth = 2, relief = tk.SUNKEN, selectmode = tk.SINGLE)
listbox_serial.pack( side = tk.LEFT, fill = tk.BOTH, expand = True)
listbox_serial.bind( '<<ListboxSelect>>', click_serial)
# listbox_serial.bind( '<Double-1>', double_click_serial)

frame_adddel = tk.Frame( frame_serial)
frame_adddel.pack( side = tk.RIGHT)

button_add = tk.Button( frame_adddel, height = 2, text = '+', command = click_add)
button_add.pack( side = tk.TOP, padx = 4, pady = 6)
button_del = tk.Button( frame_adddel, height = 2, text = '-', command = click_del)
button_del.pack( side = tk.TOP, padx = 4, pady = 6)

frame_csv = tk.Frame( window_main)
frame_csv.pack( side = tk.TOP, fill = tk.X, padx = 4, pady = 4)

tk.Label( frame_csv, anchor = tk.W, text = "File:").pack( side = tk.LEFT)

entry_csv = tk.Entry( frame_csv, borderwidth = 2, relief = tk.SUNKEN, textvariable = var_file, state = 'readonly')
entry_csv.pack( side = tk.LEFT, fill = tk.X, expand = True, padx = 4, pady = 4)

botton_browse = tk.Button( frame_csv, text = 'Browse', command = click_browse)
botton_browse.pack( side = tk.RIGHT, padx = 4, pady = 4)

frame_basic = tk.Frame( window_main)
frame_basic.pack( side = tk.TOP, fill = tk.X, padx = 4, pady = 4)

tk.Label( frame_basic, anchor = tk.W, text = "Name:").pack( side = tk.LEFT)

entry_name = tk.Entry( frame_basic, textvariable = var_name)
entry_name.pack( side = tk.LEFT, fill = tk.X, expand = True)

radio_horizontal = tk.Radiobutton( frame_basic, text = 'Horizontal', variable = var_dir, value = DIR_HORIZONTAL)
radio_horizontal.pack( side = tk.RIGHT, padx = 4)

radio_vertical = tk.Radiobutton( frame_basic, text = 'Vertical', variable = var_dir, value = DIR_VERTICAL)
radio_vertical.pack( side = tk.RIGHT, padx = 4)

frame_coordinates = tk.Frame( window_main)
frame_coordinates.pack( side = tk.TOP, fill = tk.X, padx = 4, pady = 4)

tk.Label( frame_coordinates, anchor = tk.W, text = "FROM:").pack( side = tk.LEFT)
entry_from = tk.Entry( frame_coordinates, textvariable = var_from, width = 3)
entry_from.pack( side = tk.LEFT)

tk.Label( frame_coordinates, anchor = tk.W, text = " TO:").pack( side = tk.LEFT)
entry_to = tk.Entry( frame_coordinates, textvariable = var_to, width = 3)
entry_to.pack( side = tk.LEFT)

tk.Label( frame_coordinates, anchor = tk.W, text = " X:").pack( side = tk.LEFT)
entry_x = tk.Entry( frame_coordinates, textvariable = var_x, width = 7)
entry_x.pack( side = tk.LEFT)

tk.Label( frame_coordinates, anchor = tk.W, text = " Y:").pack( side = tk.LEFT)
entry_y = tk.Entry( frame_coordinates, textvariable = var_y, width = 7)
entry_y.pack( side = tk.LEFT)

tk.Label( frame_coordinates, anchor = tk.W, text = " Sheet:").pack( side = tk.LEFT)
entry_sheet = tk.Entry( frame_coordinates, textvariable = var_sheet, width = 8)
if not openpyxl_loaded:
    entry_sheet.config( state = tk.DISABLED)
entry_sheet.pack( side = tk.LEFT)

check_hide = tk.Checkbutton( frame_coordinates, anchor = tk.W, text = "Hide", variable = var_hide)
check_hide.pack( side = tk.RIGHT)

frame_trendline = tk.Frame( window_main)
frame_trendline.pack( side = tk.TOP, fill = tk.X, padx = 4, pady = 4)

tk.Label( frame_trendline, anchor = tk.W, text = "Trendline:").pack( side = tk.LEFT)
combobox_tltype = ttk.Combobox( frame_trendline, textvariable = var_tltype, state = 'readonly', values = ( TL_NONE, TL_POLY, TL_AVG, TL_VERTICAL, TL_HORIZONTAL))
combobox_tltype.current( 0)
combobox_tltype.pack( side = tk.LEFT)

tk.Label( frame_trendline, anchor = tk.W, text = " N:").pack( side = tk.LEFT)
entry_n = tk.Entry( frame_trendline, textvariable = var_tln, width = 4)
entry_n.pack( side = tk.LEFT)

frame_filter = tk.Frame( window_main)
frame_filter.pack( side = tk.TOP, fill = tk.BOTH, expand = True, padx = 4, pady = 4)

tk.Label( frame_filter, anchor = tk.W, text = "Filters: (python expression, e.g. \"${A}=='type1' and int(${2})<9\")").pack( side = tk.TOP, fill = tk.X)
text_filter = tkst.ScrolledText( frame_filter, height = 10, borderwidth = 2, relief = tk.SUNKEN, highlightthickness = 0)
text_filter.pack( side = tk.BOTTOM, fill = tk.BOTH, expand = True)

button_update = tk.Button( window_main, height = 2, text = 'Update', command = click_update)
button_update.pack( side = tk.BOTTOM, fill = tk.X, padx = 18, pady = 6)

if DND_ENABLED:
    window_main.drop_target_register( DND_FILES)
    window_main.dnd_bind( '<<DropEnter>>', on_dnd_enter)
    window_main.dnd_bind( '<<Drop>>', on_dnd_drop)

# Run:

window_main.mainloop()


# End of 'draw_scatter.py' 

